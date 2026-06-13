#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Theme Clustering Tool

Clusters sub-themes into a consolidated theme pool using an LLM.
Works in two phases:
1. Generate an initial theme pool from the first N sub-themes.
2. Assign each sub-theme to an existing theme or create a new one.

Python 3.11+
"""

import os
import argparse
from tqdm import tqdm
import pandas as pd
from openai import OpenAI
import httpx
import openpyxl


def create_client(base_url: str, api_key: str) -> OpenAI:
    """Create an OpenAI-compatible client."""
    return OpenAI(
        base_url=base_url,
        api_key=api_key,
        http_client=httpx.Client(
            base_url=base_url,
            follow_redirects=True,
        ),
    )


def load_prompt(prompt_path: str | None, default: str) -> str:
    """Load a prompt from file if provided, otherwise use the default."""
    if prompt_path and os.path.isfile(prompt_path):
        with open(prompt_path, "r", encoding="utf-8") as f:
            return f.read().strip()
    return default


DEFAULT_INIT_PROMPT = (
    "You will receive a list of sub-themes. Your task is to merge those that are "
    "paraphrases or near-duplicates, and output 3-6 base themes.\n"
    "Requirements: each theme name must be 4-6 characters.\n"
    "Output format: themes separated by Chinese semicolons (；), no other text."
)

DEFAULT_ASSIGN_PROMPT = (
    "You will receive the current theme pool (separated by ；) and a new sub-theme.\n"
    "Task: determine if the new sub-theme belongs to any existing theme "
    "(same or similar meaning).\n"
    "- If yes, reply with the exact existing theme name.\n"
    "- If no, create a new theme name (4-6 characters) and reply with it only."
)


def call_llm(client: OpenAI, model: str, system_content: str,
             user_msg: str, temperature: float = 0.1) -> str:
    """Send a message to the LLM and return the response."""
    completion = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system_content},
            {"role": "user", "content": user_msg},
        ],
        temperature=temperature,
    )
    return completion.choices[0].message.content


def generate_initial_pool(client: OpenAI, model: str, init_prompt: str,
                          sub_themes: list[str], pool_size: int,
                          separator: str, temperature: float) -> list[str]:
    """Phase 1: Generate the initial theme pool from the first N sub-themes."""
    num_for_pool = min(pool_size, len(sub_themes))
    first_n = sub_themes[:num_for_pool]
    formatted = "\n".join(f"{i + 1}. {s}" for i, s in enumerate(first_n))
    user_msg = f"Cluster the following sub-themes:\n{formatted}"

    response = call_llm(client, model, init_prompt, user_msg, temperature)
    pool = [p.strip() for p in response.split(separator) if p.strip()]
    return pool


def assign_themes(client: OpenAI, model: str, assign_prompt: str,
                  sub_themes: list[str], pool: list[str], sheet,
                  output_col: int, separator: str, save_interval: int,
                  workbook, file_path: str, temperature: float) -> list[str] | None:
    """Phase 2: Assign each sub-theme to a theme, expanding the pool as needed."""
    for idx, sub_theme in enumerate(tqdm(sub_themes, desc="Assigning themes", ncols=100)):
        try:
            current_pool_str = separator.join(pool)
            user_msg = (
                f"Current theme pool: {current_pool_str}\n\n"
                f"New sub-theme: {sub_theme}"
            )
            response = call_llm(client, model, assign_prompt, user_msg, temperature)
            theme = response.strip()

            if not theme:
                theme = "unknown"

            if theme not in pool:
                pool.append(theme)
                print(f"New theme added to pool: {theme}")

            row = idx + 2
            sheet.cell(row=row, column=output_col).value = theme

        except Exception as e:
            print(f"Row {idx + 1} error: {e}")
            sheet.cell(row=idx + 2, column=output_col).value = "error"

        finally:
            if (idx + 1) % save_interval == 0:
                workbook.save(file_path)

    return pool


def process(file_path: str, client: OpenAI, model: str,
            init_prompt: str, assign_prompt: str,
            input_column: str, output_col: int, pool_size: int,
            separator: str, save_interval: int,
            temperature: float) -> None:
    """Main processing pipeline."""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    df = pd.read_excel(file_path)

    if input_column not in df.columns:
        raise ValueError(
            f"Column '{input_column}' not found. "
            f"Available columns: {list(df.columns)}"
        )

    contents = [
        str(item).replace("\\", "").strip()
        for item in df[input_column]
        if pd.notna(item) and str(item).strip()
    ]

    if not contents:
        print("No data found.")
        return

    # Phase 1: Generate initial theme pool
    print("Phase 1: Generating initial theme pool...")
    try:
        pool = generate_initial_pool(
            client, model, init_prompt, contents, pool_size, separator, temperature
        )
        print(f"Initial pool ({len(pool)} themes): {separator.join(pool)}")
    except Exception as e:
        print(f"Initial clustering failed: {e}")
        pool = []

    # Phase 2: Assign themes to all sub-themes
    print("Phase 2: Assigning themes to sub-themes...")
    pool = assign_themes(
        client, model, assign_prompt, contents, pool, sheet,
        output_col, separator, save_interval, workbook, file_path, temperature
    )

    workbook.save(file_path)
    print(f"Done. Final pool ({len(pool)} themes): {separator.join(pool)}")


def main():
    parser = argparse.ArgumentParser(
        description="LLM-based sub-theme clustering tool"
    )
    parser.add_argument(
        "--file", type=str, required=True,
        help="Path to the input Excel file"
    )
    parser.add_argument(
        "--column", type=str, default="sub-theme",
        help="Name of the sub-theme column (default: sub-theme)"
    )
    parser.add_argument(
        "--output-col", type=int, default=3,
        help="1-based column index to write assigned themes (default: 3)"
    )
    parser.add_argument(
        "--pool-size", type=int, default=53,
        help="Number of initial sub-themes used to generate the base pool (default: 53)"
    )
    parser.add_argument(
        "--base-url", type=str,
        default=os.getenv("LLM_BASE_URL", "https://api.openai.com/v1"),
        help="LLM API base URL (or set LLM_BASE_URL env var)"
    )
    parser.add_argument(
        "--api-key", type=str,
        default=os.getenv("LLM_API_KEY", ""),
        help="LLM API key (or set LLM_API_KEY env var)"
    )
    parser.add_argument(
        "--model", type=str,
        default=os.getenv("LLM_MODEL", "deepseek-chat"),
        help="Model name (default: deepseek-chat)"
    )
    parser.add_argument(
        "--init-prompt-file", type=str, default=None,
        help="Path to custom init clustering prompt file"
    )
    parser.add_argument(
        "--assign-prompt-file", type=str, default=None,
        help="Path to custom assignment prompt file"
    )
    parser.add_argument(
        "--separator", type=str, default="；",
        help="Separator for theme lists (default: ；)"
    )
    parser.add_argument(
        "--save-interval", type=int, default=20,
        help="Save workbook every N rows (default: 20)"
    )
    parser.add_argument(
        "--temperature", type=float, default=0.1,
        help="Generation temperature (default: 0.1)"
    )
    args = parser.parse_args()

    if not args.api_key:
        raise SystemExit(
            "Provide an API key via --api-key or the LLM_API_KEY env var."
        )

    client = create_client(args.base_url, args.api_key)
    init_prompt = load_prompt(args.init_prompt_file, DEFAULT_INIT_PROMPT)
    assign_prompt = load_prompt(args.assign_prompt_file, DEFAULT_ASSIGN_PROMPT)

    process(
        file_path=args.file,
        client=client,
        model=args.model,
        init_prompt=init_prompt,
        assign_prompt=assign_prompt,
        input_column=args.column,
        output_col=args.output_col,
        pool_size=args.pool_size,
        separator=args.separator,
        save_interval=args.save_interval,
        temperature=args.temperature,
    )


if __name__ == "__main__":
    main()
